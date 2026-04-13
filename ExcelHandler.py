from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.cell import Cell, MergedCell
from typing import List, Callable
from utils import get_formatted_value
from Config import Config

class CorrespondingData:
    def __init__(self, col_name: str, data: List[str]):
        self.col_name = col_name
        self.data = data
    
    def get_data_tuple(self):
        return (self.col_name, self.data)
    
''' 
    Takes name of Excel file and worksheet to use from that file,
    performs read and write operations on that Excel file
'''
class ExcelHandler:
    def __init__(self, filename: str, config: Config):
        print("Initializing Excel Workbook", filename)
        self.filename = filename
        self.wb = load_workbook(filename=self.filename, data_only=True)
        self.ws = self.wb[config.get_excel_config('worksheet_name')]
        self.no_ws_err = f"No Worksheet found for {self.filename}"
        self.config = config

    ''' Adds data to cell in string format {col_letter}{row_number} '''
    def add_data_to_cell(self, data: str, cell: str):
        print("Adding data", data, "to cell", cell)
        if self.ws is None:
            raise Exception(self.no_ws_err)
        self.ws[cell] = data

    def save_file_changes(self):
        print("\nSaving changes to Excel Workbook", self.filename, "\n")
        self.wb.save(self.filename)

    '''
        Purpose: Adds lists of data to excel meant to correspond with an ID column
    '''
    def add_corresponding_data(self, row_num: int, idx: int, data_lists: List[CorrespondingData]):
        if self.ws is None or len(data_lists) <= 0:
            raise Exception(self.no_ws_err)
        for corresponding_data in data_lists:
            (col_name, data) = corresponding_data.get_data_tuple()
            cell = f"{col_name}{row_num}"
            self.add_data_to_cell(data=data[idx], cell=cell)

    '''
        Using row #, returns a dictionary that can be used with Mail Merge API
        The returned dictionary is based off merge_name_map in utils.py
    '''
    def get_merge_data_from_row(self, row_num: int) -> dict[str, str | float | dict]:
        if self.ws is None:
            raise Exception(self.no_ws_err)
        row_values = [cell.value for cell in self.ws[row_num]]
        filtered = { k:v for k,v in self.config.excel_config.items() if v.isalpha() and len(v) == 1 }
        keys = filtered.keys()
        values = filtered.values()
        column_idxs = [column_index_from_string(v) - 1 for v in values]
        zipped = zip(keys, column_idxs)
        return { k: get_formatted_value(key=k, value=str(row_values[v])) for (k,v) in zipped }

    '''
        Purpose: Perform operations on specific rows filtered by IDs
        callback: Function to be called for each matched row, gives row info
        id_col_letter: The column letter (like "G") that has the IDs
        ids: List of IDs to match
        ids_asc: True if IDs in the id_col_letter are logically ascending
    '''
    def iterate_rows_by_ids_bind(
        self, 
        callback: Callable[[tuple[Cell | MergedCell, ...], str, int, int], None],
        id_col_letter: str,
        ids: List[str],
        ids_asc: bool = True
    ):
        if self.ws is None:
            raise Exception(self.no_ws_err)
        col_num = column_index_from_string(id_col_letter) - 1
        row_num = 2
        errors = ""
        for idx, id in enumerate(ids):
            for row in self.ws.iter_rows(min_row=row_num if ids_asc else 2):
                if row[col_num].value is None:
                    print("ID", id, "not found in spreadsheet.")
                    errors += f"ID {id} not found in spreadsheet."
                    break
                row_num = row[col_num].row
                if row[col_num].value == id:
                    callback(row, id, idx, row_num if row_num is not None else 2)
                    break     
        if errors:
            raise Exception(errors)   